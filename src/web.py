import base64
import json
import os
import re
import sys
from concurrent.futures.thread import ThreadPoolExecutor
from hashlib import sha256
from urllib.parse import unquote

import asyncio
import chardet
import requests
from aiohttp import web
from ldap3 import Server, Connection, ALL, NTLM
from openpyxl import load_workbook
from requests_ntlm import HttpNtlmAuth

from mailadapter import AUTOFAQ_SERVICE_HOST, send_to_user
from models.db import Database
from web import Incident
from web.queries import get, add_phone, update_incident, get_incident
from web.settings import *


def laps(server, user, password, computer_name):
    server = Server(server, get_info=ALL)
    conn = Connection(server=server, user=user, password=password, authentication=NTLM)
    conn.bind()
    conn.search(
        "dc=Dixy,dc=local",
        f"(&(objectclass=computer)(name={computer_name}))",
        attributes=["ms-Mcs-AdmPwd"],
    )
    if conn.entries:
        return str(conn.entries[0]["ms-Mcs-AdmPwd"])
    else:
        return None


async def handle_laps(request):
    text = await request.text()
    try:
        body = json.loads(text or "{}")
    except json.JSONDecodeError:
        return web.Response(status=400)
    server = body.get("server", None)
    user = body.get("user", None)
    password = body.get("password", None)
    computer_name = body.get("computerName", None)
    if server and user and password and computer_name is not None:
        text = (
            laps(
                server=server, user=user, password=password, computer_name=computer_name
            )
            or "0"
        )
    else:
        text = "0"
    return web.Response(text=text)


def bio(user, password, computer_name):
    response = requests.get(
        f"http://10.0.17.148/ReportServer/Pages/ReportViewer.aspx?%2fBiolink%2fBioTime+Manager+pwd&rs:Command=Value&Login={computer_name}",
        auth=HttpNtlmAuth(username=user, password=password),
    )
    result = re.findall(
        '<div style="width:28.09mm;min0-width: 28.09mm;">(\d+)</div>', response.text
    )
    if result:
        return result[0]
    else:
        return None


async def handle_bio(request):
    text = await request.text()
    try:
        body = json.loads(text or "{}")
    except json.JSONDecodeError:
        return web.Response(status=400)
    user = body.get("user", None)
    password = body.get("password", None)
    shop_num = body.get("shopNum", None)
    if user and password and shop_num is not None:
        text = bio(user=user, password=password, computer_name=shop_num) or "0"
    else:
        text = "0"
    return web.Response(text=text)


def basic_auth(username, password):
    auth_str = f"{username}:{password}"
    try:
        return base64.b64encode(auth_str.encode("utf-8")).decode("utf-8")
    except:
        return None


async def handle_basic_auth(request):
    text = await request.text()
    try:
        body = json.loads(text or "{}")
    except json.JSONDecodeError:
        return web.Response(status=400)
    username = body.get("username", None)
    password = body.get("password", None)
    if username and password is not None:
        text = basic_auth(username=username, password=password)
        if text is None:
            return web.Response(status=500)
    else:
        return web.Response(status=400)
    return web.Response(text=text)


async def handle_auth_get(request):
    email = request.rel_url.query.get("email", None)
    phone = request.rel_url.query.get("phone", None)
    user_info = {"state": False}

    conn = request.app["db"]
    user = await conn.fetchrow(
        """
        SELECT * 
        FROM users 
        WHERE email = $1 OR phone = $2
    """,
        email,
        phone,
    )
    if user:
        user_info["state"] = True
        user_info["full_name"] = user["full_name"]
        user_info["position"] = user["position"]
    return web.Response(
        text=json.dumps(user_info, ensure_ascii=False, indent=4), charset="utf-8"
    )


def get_user_info(username, password):
    try:
        server = Server(os.environ["AD"], get_info=ALL)
        conn = Connection(
            server=server,
            user=f"dixy\\{username}",
            password=password,
            authentication=NTLM,
        )
        conn.bind()
        conn.search(
            "dc=Dixy,dc=local",
            f"(&(objectclass=person)(sAMAccountName={username}))",
            attributes=["title", "cn", "employeeID"],
        )
    except:
        return None
    if conn.entries:
        user = conn.entries[0]
        return {
            "user_id": str(user["employeeID"]),
            "full_name": str(user["cn"]),
            "position": str(user["title"]),
        }
    else:
        return None


async def handle_auth_post(request):
    text = await request.text()
    try:
        body = json.loads(text or "{}")
    except json.JSONDecodeError:
        return web.Response(status=400)
    username = body.get("username", None)
    password = body.get("password", None)
    email = body.get("email", None)
    phone = body.get("phone", None)
    if ((username and password) is not None) and ((email or phone) is not None):
        user = get_user_info(username=username, password=password)
        if user:
            user["email"] = email
            user["phone"] = phone
            conn = request.app["db"]
            await conn.execute(
                """
                INSERT INTO Users(user_id, email, full_name, position, phone)
                VALUES($1, $2, $3, $4, $5)
            """,
                user["user_id"],
                user["email"],
                user["full_name"],
                user["position"],
                user["phone"],
            )
        else:
            return web.Response(status=401)
    else:
        return web.Response(status=400)
    return web.Response(text=text)


async def handle_get_incidents(request):
    user_initiator = request.rel_url.query.get("userInitiator", None)
    if sha256(
        request.headers["Authorization"].encode("utf-8")
    ).hexdigest() == os.environ.get("SECRET_SHA256"):
        response = requests.get(
            f"{os.environ.get('ITIL_API_URL')}getListIncidents/?UserInitiator={user_initiator}",
            auth=(os.environ.get("ITIL_LOGIN"), os.environ.get("ITIL_PASS")),
        )
        return web.Response(text=response.content.decode("utf-8"), charset="utf-8")
    else:
        return web.Response(status=401)


async def handle_get_services(request):
    user_initiator = request.rel_url.query.get("userInitiator", None)
    if sha256(
        request.headers["Authorization"].encode("utf-8")
    ).hexdigest() == os.environ.get("SECRET_SHA256"):
        response = requests.get(
            f"{os.environ.get('ITIL_API_URL')}getListServices/?UserInitiator={user_initiator}",
            auth=(os.environ.get("ITIL_LOGIN"), os.environ.get("ITIL_PASS")),
        )
        return web.Response(text=response.text, charset="utf-8")
    else:
        return web.Response(status=401)


def parse_phone(phone: str) -> int:
    phone = "".join(re.findall(r"\d+", phone))
    if len(phone) < 10:
        raise ValueError
    try:
        phone = int(phone[len(phone) - 10 :])
    except:
        raise ValueError
    return phone


def get_extra_user(phone):
    contractors_file = DATA_DIR / "users.xlsx"
    partners_file = DATA_DIR / "partners.xlsx"
    user = get_user_xlsx(partners_file, phone)
    if user is not None:
        user["position"] = "Партнёр"
        return user
    user = get_user_xlsx(contractors_file, phone)
    if user is not None:
        user["position"] = "Подрядчик"
        return user
    return None


def get_user_xlsx(file, phone):
    if type(phone) == str:
        try:
            phone = parse_phone(phone)
        except ValueError:
            return None

    try:
        wb = load_workbook(file)
    except:
        logging.error(f"Failed to open {file}")
        return None
    ws = wb.active
    for row in ws.iter_rows():
        try:
            if parse_phone(str(row[1].value)) == phone:
                return {
                    "phone": phone,
                    "state": "true",
                    "user_num": "",
                    "fio": row[0].value,
                    "shop_num": row[2].value,
                    # 'position': 'Подрядчик',
                    "login": "",
                }
        except ValueError:
            continue
    return None


async def get_user(phone):
    file = DATA_DIR / "users.txt"

    try:
        phone = parse_phone(phone)
    except:
        return {"state": False, "comment": "Некорректный номер телефона"}

    try:
        with open(file, encoding="utf-8") as data:
            lines = data.read().splitlines()
    except:
        return {"state": False, "comment": "Не нашел файл users.txt"}

    if len(lines) == 0:
        return {"state": False, "comment": "Не получил из файла данных"}

    users = []
    for line in lines:
        _temp = line.split(",")
        if len(_temp) < 6:
            continue

        _phone = (
            _temp[0]
            .strip()
            .replace(" ", "")
            .replace("-", "")
            .replace("_", "")
            .replace("/", "")
            .replace(".", "")
            .replace("!", "")
            .replace("&", "")
            .replace("(", "")
            .replace(")", "")
        )
        _num = _temp[1].strip()
        _fio = _temp[2].strip()
        _shop = _temp[3].strip()
        _position = _temp[4].strip()
        _login = _temp[5].strip()

        # if len(_phone) < 10:
        #     continue

        try:
            _phone = int(_phone[len(_phone) - 10 :])
        except:
            _phone = None

        users.append(
            {
                "phone": _phone,
                "state": "true",
                "user_num": _num,
                "fio": _fio,
                "shop_num": _shop,
                "position": _position,
                "login": _login,
            }
        )

    if len(users) == 0:
        return {
            "state": False,
            "comment": "Не смог загрузить данные из файла, проверьте формат и структуру данных",
        }

    matches = [i for i in users if i["phone"] == phone]

    if len(matches) == 0:
        res = await get(phone=str(phone))
        if res:
            for user in users:
                if user["user_num"] == res["user_id"]:
                    user["phone"] = phone
                    matches.append(user)

    extra = get_extra_user(phone)
    if extra:
        matches.append(extra)

    if len(matches) == 0:
        return {"state": False, "comment": "Не нашел номер телефона в таблице"}

    return matches[0]


async def handle_get_user(request):
    phone = request.rel_url.query.get("phone", None)
    if sha256(
        request.headers["Authorization"].encode("utf-8")
    ).hexdigest() == os.environ.get("SECRET_SHA256"):
        if phone is not None:
            service = await get_user(str(phone).strip())
        else:
            service = {"state": False}
        return web.Response(
            text=json.dumps(service, ensure_ascii=False), charset="utf-8"
        )
    else:
        return web.Response(status=401)


async def handle_add_user_phone(request):
    if request.headers["Authorization"] == os.environ.get("SECRET"):
        text = await request.text()
        try:
            body = json.loads(text or "{}")
        except json.JSONDecodeError:
            return web.Response(status=400)
        phone = body.get("phone", None)
        employee_id = body.get("employeeId", None)
        if (phone and employee_id) is not None:
            try:
                phone = parse_phone(str(phone).strip())
                await add_phone(user_id=employee_id, phone=str(phone))
            except ValueError:
                web.Response(status=400)
        else:
            return web.Response(status=400)
    else:
        return web.Response(status=401)

    return web.Response(status=200)


async def handle_send_files_to_itil(request):
    text = await request.text()
    try:
        body = json.loads(text or "{}")
    except json.JSONDecodeError:
        return web.Response(status=400)
    files_ids = body.get("files_ids", None)
    uid = body.get("uid", None)
    if files_ids and uid:
        for file_id in json.loads(files_ids):
            url = f"http://{AUTOFAQ_SERVICE_HOST}/api/files/"
            r = requests.get(url=url + file_id)
            d = r.headers["content-disposition"]
            file_name = unquote(re.findall('filename="(.+)"', d)[0])
            base64_image = base64.b64encode(r.content).decode("utf-8")
            j = json.dumps({"UID": uid, "NameFile": file_name, "Data": base64_image})
            response = requests.post(
                f"{os.environ.get('ITIL_API_URL')}addFileToIncident",
                auth=(os.environ.get("ITIL_LOGIN"), os.environ.get("ITIL_PASS")),
                data=j,
            )
        return web.Response(status=200)

    return web.Response(status=400)


async def handle_create_incident(request):
    text = await request.text()
    try:
        body = json.loads(text or "{}")
    except json.JSONDecodeError:
        return web.Response(status=400)
    user_id = body.get("UserID", None)
    if user_id:
        data = json.dumps(body)
        response = requests.post(
            f"{os.environ.get('ITIL_API_URL')}addNewIncident",
            auth=(os.environ.get("ITIL_LOGIN"), os.environ.get("ITIL_PASS")),
            data=data,
        )
        if response.status_code == 200:
            response_json = response.json()
            uid = response_json.get("UID")
            await update_incident(
                Incident(
                    incident_uid=uid, user_tg_id=user_id, status="Зарегистрировано"
                )
            )
            return web.Response(status=200, text=response.text)

    return web.Response(status=400)


async def itil_feedback():
    while True:
        try:
            response = requests.get(
                f"{os.environ.get('ITIL_API_URL')}getListIncidents",
                auth=(os.environ.get("ITIL_LOGIN"), os.environ.get("ITIL_PASS")),
            )
            if response.status_code == 200:
                response_json = response.json()
                for incident in response_json:
                    status = incident.get("Status", {}).get("Name", None)
                    uid = incident.get("UID", None)
                    if status and uid:
                        res = await get_incident(incident_uid=uid)
                        if res:
                            if res["status"] != status:
                                if status == "На уточнении":
                                    send_to_user(
                                        message=f"Обращение {incident['Number']}: {status}",
                                        user_id=res["user_tg_id"],
                                    )

                                if status == "Выполнено. Требует подтверждения":
                                    send_to_user(
                                        message=f"Обращение {incident['Number']}: {status}",
                                        user_id=res["user_tg_id"],
                                    )
                                await update_incident(
                                    Incident(
                                        user_tg_id=res["user_tg_id"],
                                        incident_uid=uid,
                                        status=status,
                                    )
                                )
        except:
            logging.exception()

        await asyncio.sleep(300)


async def init_app():
    app = web.Application()
    # db = await Database.get_connection_pool()
    # app["db"] = db
    app.add_routes([web.post("/laps", handle_laps)])
    app.add_routes([web.post("/bio", handle_bio)])
    app.add_routes([web.post("/basic-auth", handle_basic_auth)])
    app.add_routes([web.post("/auth", handle_auth_post)])
    app.add_routes([web.get("/auth", handle_auth_get)])
    app.add_routes([web.get("/itil-get-incidents", handle_get_incidents)])
    app.add_routes([web.get("/itil-get-services", handle_get_services)])
    app.add_routes([web.get("/user", handle_get_user)])
    app.add_routes([web.post("/add-user-phone", handle_add_user_phone)])
    app.add_routes([web.post("/send-file-to-itil", handle_send_files_to_itil)])
    app.add_routes([web.post("/itil-create-incident", handle_create_incident)])

    loop = asyncio.get_event_loop()
    app["itil_feedback_thread"] = loop.create_task(itil_feedback())
    return app


if __name__ == "__main__":
    web.run_app(init_app(), host="0.0.0.0", port=4444)
