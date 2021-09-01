import json

from openpyxl import load_workbook

from settings import DATA_DIR
from utils import parse_phone
from aiohttp import web

from web.queries import add_phone, get_user_by_phone


def get_extra_user(phone):
    contractors_file = DATA_DIR / "users.xlsx"
    partners_file = DATA_DIR / "partners.xlsx"
    user = get_user_xlsx(partners_file, phone)
    if user is not None:
        user["position"] = "Партнёр"
        return user
    user = get_user_xlsx(contractors_file, phone)
    if user is not None:
        user["position"] = "Подрядчик"
        return user
    return None


def get_user_xlsx(file, phone):
    if type(phone) == str:
        try:
            phone = parse_phone(phone)
        except ValueError:
            return None

    try:
        wb = load_workbook(file)
    except:
        logging.error(f"Failed to open {file}")
        return None
    ws = wb.active
    for row in ws.iter_rows():
        try:
            if parse_phone(str(row[1].value)) == phone:
                return {
                    "phone": phone,
                    "state": "true",
                    "user_num": "",
                    "fio": row[0].value,
                    "shop_num": row[2].value,
                    # 'position': 'Подрядчик',
                    "login": "",
                }
        except ValueError:
            continue
    return None


async def get_user(phone):
    file = DATA_DIR / "users.txt"

    try:
        phone = parse_phone(phone)
    except:
        return {"state": False, "comment": "Некорректный номер телефона"}

    try:
        with open(file, encoding="utf-8") as data:
            lines = data.read().splitlines()
    except:
        return {"state": False, "comment": "Не нашел файл users.txt"}

    if len(lines) == 0:
        return {"state": False, "comment": "Не получил из файла данных"}

    users = []
    for line in lines:
        _temp = line.split(",")
        if len(_temp) < 6:
            continue

        _phone = (
            _temp[0]
            .strip()
            .replace(" ", "")
            .replace("-", "")
            .replace("_", "")
            .replace("/", "")
            .replace(".", "")
            .replace("!", "")
            .replace("&", "")
            .replace("(", "")
            .replace(")", "")
        )
        _num = _temp[1].strip()
        _fio = _temp[2].strip()
        _shop = _temp[3].strip()
        _position = _temp[4].strip()
        _login = _temp[5].strip()

        # if len(_phone) < 10:
        #     continue

        try:
            _phone = int(_phone[len(_phone) - 10 :])
        except:
            _phone = None

        users.append(
            {
                "phone": _phone,
                "state": "true",
                "user_num": _num,
                "fio": _fio,
                "shop_num": _shop,
                "position": _position,
                "login": _login,
            }
        )

    if len(users) == 0:
        return {
            "state": False,
            "comment": "Не смог загрузить данные из файла, проверьте формат и структуру данных",
        }

    matches = [i for i in users if i["phone"] == phone]

    if len(matches) == 0:
        res = await get_user_by_phone(phone=str(phone))
        if res:
            for user in users:
                if user["user_num"] == res["user_id"]:
                    user["phone"] = phone
                    matches.append(user)

    extra = get_extra_user(phone)
    if extra:
        matches.append(extra)

    if len(matches) == 0:
        return {"state": False, "comment": "Не нашел номер телефона в таблице"}

    return matches[0]


class UserView(web.View):
    async def get(self):
        phone = self.request.query.get("phone")
        if phone:
            service = await get_user(str(phone).strip())
        else:
            service = {"state": False}
        return web.Response(
            text=json.dumps(service, ensure_ascii=False), charset="utf-8"
        )


async def handle_add_user_phone(request):
    if request.headers["Authorization"] == os.environ.get_user_by_phone("SECRET"):
        text = await request.text()
        try:
            body = json.loads(text or "{}")
        except json.JSONDecodeError:
            return web.Response(status=400)
        phone = body.get_user_by_phone("phone", None)
        employee_id = body.get_user_by_phone("employeeId", None)
        if (phone and employee_id) is not None:
            try:
                phone = parse_phone(str(phone).strip())
                await add_phone(user_id=employee_id, phone=str(phone))
            except ValueError:
                web.Response(status=400)
        else:
            return web.Response(status=400)
    else:
        return web.Response(status=401)

    return web.Response(status=200)
